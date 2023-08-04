# main.py

import gc
import pdf2image
from PIL import Image
import pytesseract
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers

def process_pdf(file_path):
    # DPI for the conversion, lower this to decrease memory usage but keep in mind that it could affect OCR accuracy
    dpi = 200

    # Set of valid attributes
    valid_attributes = set([
        "Creative",
        "Original",
        "Curious",
        "Non-Conforming",
        "Deliberative",
        "Logical",
        "Systematic",
        "Impartial",
        "Detailed and Reliable",
        "Organized",
        "Detail-Oriented",
        "Dependable",
        "Conceptual",
        "Practical",
        "Extraverted",
        "Gregarious",
        "Engaging",
        "Adventurous",
        "Tough",
        "Feisty",
        "Critical",
        "Direct",
        "Nurturing",
        "Helpful",
        "Empathetic",
        "Person-Oriented",
        "Leadership",
        "Taking Charge",
        "Inspiring",
        "Demanding",
        "Humorous",
        "Composed",
        "Calm",
        "Confident",
        "Poised",
        "Autonomous",
        "Independent",
        "Self-Accountable",
        "Internally Motivated",
        "Flexible",
        "Adaptable",
        "Agile",
        "Growth-Seeking",
        "Determined",
        "Persistent",
        "Driven",
        "Proactive",
        "Humble",
        "Receptive to Criticism",
        "Open-Minded",
        "Modest",
        "Energetic",
        "Status-Seeking",
    ])

    # Dictionary to hold the attributes
    attributes = {}

    # Variable to hold all text
    all_text = ""

    # Get a list of pages
    pages = pdf2image.convert_from_path(file_path, dpi)

    # Iterate over all the pages in the PDF and convert each to image
    for i, image in enumerate(pages):
        text = pytesseract.image_to_string(image)
        all_text += text + "\n"
        matches = re.findall(r"(.+)\s+(\d+)%|(\d+)%\s+(.+)", text)
        for match in matches:
            if match[0]:
                attribute = match[0].strip()
                percentage = int(match[1]) 
            else:
                attribute = match[3].strip()
                percentage = int(match[2]) 

            if percentage > 100 and percentage % 10 == 7:
                percentage = int(percentage / 10) 
            if attribute in valid_attributes and attribute not in attributes:   
                attributes[attribute] = str(percentage) + "%"

        # Delete the image and text variable to free up memory
        del image
        del text

        # Run garbage collector
        gc.collect()

    # Write OCR results to a text file
    with open("ocrResult.txt", "w") as text_file:
        text_file.write(all_text)

    # Create a DataFrame from the dictionary
    df = pd.DataFrame(list(attributes.items()), columns=["Attribute", "Score"])

    # Write the DataFrame to an Excel file
    df.to_excel("attributes.xlsx", index=False)

    # Open the workbook
    book = load_workbook("attributes.xlsx")

    # Get the writer's sheet
    ws = book.active

    # Apply number format
    for row in ws['B2:B'+str(len(df)+1)]:
        for cell in row:
            cell.number_format = numbers.FORMAT_PERCENTAGE_00

    # Save the changes
    book.save("attributes.xlsx")

    print("Final attributes scores:")
    for k, v in attributes.items():
        print(f"{k}: {v}")

    return attributes
